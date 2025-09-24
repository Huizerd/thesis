import cv2
import numpy as np
import sys
import openpyxl

def hsv_to_hex(hsv):
    bgr = cv2.cvtColor(np.array([[hsv]], dtype=np.uint8), cv2.COLOR_HSV2BGR)[0][0]
    return '#{:02x}{:02x}{:02x}'.format(int(bgr[2]), int(bgr[1]), int(bgr[0]))

def generate_colors(v1, s1, h1, v2, s2, hdiff):

    h2 = (h1 + hdiff) % 1
    print(v1, s1, h1, v2, s2, h2)
    hsv1 = np.array([int(h1 * 255), int(s1 * 255), int(v1 * 255)], dtype=np.uint8)
    hsv2 = np.array([int(h2 * 255), int(s2 * 255), int(v2 * 255)], dtype=np.uint8)

    # hsv to hex conversion 
    hex1 = hsv_to_hex(hsv1)
    hex2 = hsv_to_hex(hsv2)

    return hex1, hex2

if __name__ == '__main__':
    if len(sys.argv) != 7:
        print("Usage: python color_selection.py <v1> <s1> <h1> <v2> <s2> <hue_difference>")
        sys.exit(1)

    # Read command-line arguments
    v1 = float(sys.argv[1])
    s1 = float(sys.argv[2])
    h1 = float(sys.argv[3])
    v2 = float(sys.argv[4])
    s2 = float(sys.argv[5])
    hdiff = float(sys.argv[6])

    # Write color combinations to an Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    i = 0
    dH = 0.01
    while h1 + dH <= 1:
        hex1, hex2 = generate_colors(v1, s1, h1, v2, s2, hdiff)

        sheet['A{}'.format(i+1)] = i+1
        sheet['B{}'.format(i+1)] = hex1
        sheet['C{}'.format(i+1)] = hex2

        # Set cell background colors
        fill1 = openpyxl.styles.PatternFill(start_color=hex1[1:], fill_type="solid")
        fill2 = openpyxl.styles.PatternFill(start_color=hex2[1:], fill_type="solid")
        sheet['B{}'.format(i+1)].fill = fill1
        sheet['C{}'.format(i+1)].fill = fill2

        h1 += dH
        i += 1

    workbook.save(filename="color_combinations.xlsx")
    print("Color combinations saved to color_combinations.xlsx")
